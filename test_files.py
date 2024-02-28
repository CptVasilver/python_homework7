from conftest import ARCHIVE_PATH
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl import load_workbook
import csv


def test_pdf_file():
    with ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('Python Testing with Pytest (Brian Okken).pdf') as pdf_file:
            reader = PdfReader(pdf_file)

            assert len(reader.pages) == 256
            assert 'Simple, Rapid, Effective, and Scalable' in reader.pages[1].extract_text()


def test_csv_file():
    with ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('example.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))

            assert csvreader[1][0] == 'John Doe'


def test_xlsx_file():
    with ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert workbook.sheetnames[0] == "Sheet1"
            assert sheet.cell(row=14, column=2).value == "Sherron"
